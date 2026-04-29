#!/usr/bin/env python3
import json, os
from datetime import datetime
try:
    from zoneinfo import ZoneInfo
except Exception:
    ZoneInfo = None
from openpyxl import Workbook
from openpyxl.styles import Font

def load_rows(data):
    if isinstance(data, list):
        return data
    if isinstance(data, dict):
        vals = list(data.values())
        if vals and all(isinstance(v, dict) for v in vals):
            return vals
        return [data]
    raise ValueError("Unsupported JSON structure")

def headers_union_preserve(rows):
    if not rows:
        raise ValueError("Empty JSON after parsing")
    headers = list(rows[0].keys())
    seen = set(headers)
    for r in rows[1:]:
        for k in r.keys():
            if k not in seen:
                headers.append(k)
                seen.add(k)
    return headers

def autosize(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            v = "" if cell.value is None else str(cell.value)
            if len(v) > max_len:
                max_len = len(v)
        ws.column_dimensions[col_letter].width = max(10, min(80, max_len + 2))

def main():
    repo_root = os.path.dirname(os.path.abspath(__file__))
    repo_root = os.path.dirname(repo_root)  # up from automation/
    with open(os.path.join(repo_root, 'automation', 'input.json'), 'r', encoding='utf-8') as f:
        data = json.load(f)
    rows = load_rows(data)
    headers = headers_union_preserve(rows)
    wb = Workbook()
    ws = wb.active
    ws.title = 'Data'
    # Write header
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = Font(bold=True)
    # Write rows
    for r_idx, row in enumerate(rows, start=2):
        for c_idx, h in enumerate(headers, start=1):
            ws.cell(row=r_idx, column=c_idx, value=row.get(h, ""))
    ws.freeze_panes = 'A2'
    autosize(ws)
    # Build IST timestamped filename
    tz = ZoneInfo('Asia/Kolkata') if ZoneInfo else None
    now = datetime.now(tz) if tz else datetime.utcnow()
    if tz is None:
        from datetime import timedelta
        now = now + timedelta(hours=5, minutes=30)
    ip_name = os.environ.get('IP_NAME', 'GPIO')
    out_dir = os.path.join(repo_root, 'Test_Output', ip_name, 'TestPlan')
    os.makedirs(out_dir, exist_ok=True)
    fname = f"{ip_name}_TestPlan_{now.strftime('%Y%m%d')}_{now.strftime('%H%M%S')}.xlsx"
    out_path = os.path.join(out_dir, fname)
    wb.save(out_path)
    print(out_path)

if __name__ == '__main__':
    main()
