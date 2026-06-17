#!/usr/bin/env python3
import json
import os
from datetime import datetime
from zoneinfo import ZoneInfo
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

PREFERRED_COLUMNS = [
    "Index",
    "SS / Module",
    "Feature",
    "Test Case Name",
    "Test Description",
    "Speed",
    "Mode",
    "Remarks",
    "Test Steps / Procedure",
    "Impacted Registers",
    "Validation / Acceptance Criteria",
    "Gap Analysis",
]
NEW_COLUMN_HEADER = "Code Generation (Required / Not)"

LONG_WRAP_COLUMNS = set([
    "Test Description",
    "Test Steps / Procedure",
    "Validation / Acceptance Criteria",
])

def load_json(json_path: str):
    with open(json_path, 'r', encoding='utf-8') as f:
        data = json.load(f)
    if not isinstance(data, list):
        raise ValueError("json_data must be a JSON array of objects")
    for i, row in enumerate(data):
        if not isinstance(row, dict):
            raise ValueError(f"Element at index {i} is not an object")
    return data


def build_columns(data):
    all_keys = []
    seen = set()
    for row in data:
        for k in row.keys():
            if k not in seen:
                seen.add(k)
                all_keys.append(k)
    # Preserve required order first, then any extra keys not already in preferred list
    extra = [k for k in all_keys if k not in PREFERRED_COLUMNS and k != NEW_COLUMN_HEADER]
    # Final columns order
    columns = PREFERRED_COLUMNS + extra + [NEW_COLUMN_HEADER]
    return columns


def autofit(ws, columns, max_width_cap=100):
    # Estimate width per column based on content length
    for col_idx, header in enumerate(columns, start=1):
        max_len = len(str(header)) if header is not None else 0
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
            cell = row[0]
            v = cell.value
            if v is None:
                continue
            # Consider multi-line values
            if isinstance(v, str):
                parts = v.splitlines() if "\n" in v else [v]
                for p in parts:
                    max_len = max(max_len, len(p))
            else:
                max_len = max(max_len, len(str(v)))
        # Heuristic width; Excel roughly uses ~1 char per unit
        width = min(max_width_cap, max(10, max_len + 2))
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def apply_borders(ws):
    thin = Side(border_style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = border


def apply_wrap_alignment(ws, columns):
    col_map = {name: idx+1 for idx, name in enumerate(columns)}
    for name in LONG_WRAP_COLUMNS:
        if name in col_map:
            col_idx = col_map[name]
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
                for cell in row:
                    cell.alignment = Alignment(wrap_text=True, vertical='top')


def main():
    json_path = os.environ.get('JSON_PATH', 'Test_Output/PCIE/testplan.json')
    output_dir = os.environ.get('OUTPUT_DIR', 'Test_Output/PCIE')

    data = load_json(json_path)
    columns = build_columns(data)

    wb = Workbook()
    ws = wb.active
    ws.title = "TestPlan"

    # Header row
    for col_idx, header in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(vertical='center')

    # Data rows
    for r_idx, row in enumerate(data, start=2):
        for c_idx, header in enumerate(columns, start=1):
            if header == NEW_COLUMN_HEADER:
                val = None
            else:
                val = row.get(header, None)
            ws.cell(row=r_idx, column=c_idx, value=val)

    # Freeze top row
    ws.freeze_panes = 'A2'

    # Data validation for new column (entire column, all rows)
    last_col_letter = get_column_letter(len(columns))
    dv = DataValidation(type="list", formula1='"Required,Not Required"', allow_blank=True)
    dv.error = "Select a value from the list or leave blank"
    dv.errorTitle = "Invalid Selection"
    ws.add_data_validation(dv)
    dv.add(f"{last_col_letter}2:{last_col_letter}1048576")

    # Apply wrapping to long columns
    apply_wrap_alignment(ws, columns)

    # Auto-fit columns (approximation)
    autofit(ws, columns)

    # Borders for all cells in used range
    apply_borders(ws)

    # Ensure output directory exists
    os.makedirs(output_dir, exist_ok=True)

    # IST timestamp
    ist = ZoneInfo("Asia/Kolkata")
    ts = datetime.now(ist).strftime("%Y%m%d_%H%M%S")
    filename = f"testplan_{ts}.xlsx"
    out_path = os.path.join(output_dir, filename)

    wb.save(out_path)
    print(out_path)

if __name__ == '__main__':
    main()
