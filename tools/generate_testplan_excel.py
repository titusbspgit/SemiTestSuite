#!/usr/bin/env python3
# Ag-Emb-Testsuite-Excel-Generator Agent - Fallback generator
# Converts Test Plan JSON into a REAL .xlsx with required structure and formatting.
# Usage:
#   python tools/generate_testplan_excel.py --json-path Test_Input/testplan.json --output-dir Test_Output/GPIO
#
# Notes:
# - Preserves original JSON by also embedding it in a hidden sheet named "RawJSON".
# - Timestamp in filename uses IST (Asia/Kolkata): testplan_YYYYMMDD_HHMMSS.xlsx

import argparse
import json
import os
import sys
from datetime import datetime
try:
    from zoneinfo import ZoneInfo  # Python 3.9+
except Exception:
    ZoneInfo = None

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

NEW_COL_HEADER = "Code Generation (Required / Not)"
COLUMNS = [
    "Index",
    "SS / Module",
    "Feature",
    "Test Case Name",
    "Test Description",
    "Speed",
    "Mode",
    "Remarks",
    "Test Steps / Procedure",
    "Imparted Registers" if False else "Impacted Registers",  # ensure exact header
    "Validation / Acceptance Criteria",
    "Gap Analysis",
    NEW_COL_HEADER,
]

WRAP_COLS = {
    "Test Description",
    "Test Steps / Procedure",
    "Validation / Acceptance Criteria",
}

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def to_str(val):
    if val is None:
        return ""
    if isinstance(val, (dict, list)):
        return json.dumps(val, ensure_ascii=False)
    return str(val)


def ist_now_str():
    try:
        tz = ZoneInfo("Asia/Kolkata") if ZoneInfo else None
    except Exception:
        tz = None
    dt = datetime.now(tz) if tz else datetime.utcnow()
    if tz is None:
        # approximate IST by adding +5:30 if ZoneInfo unavailable
        from datetime import timedelta
        dt = dt + timedelta(hours=5, minutes=30)
    return dt.strftime("%Y%m%d_%H%M%S")


def load_json_array(path):
    if not os.path.exists(path):
        print(f"ERROR: JSON path not found: {path}", file=sys.stderr)
        sys.exit(2)
    try:
        with open(path, "r", encoding="utf-8") as f:
            data = json.load(f)
    except Exception as e:
        print(f"ERROR: Failed to parse JSON: {e}", file=sys.stderr)
        sys.exit(2)
    if not isinstance(data, list):
        print("ERROR: Top-level JSON must be an array of objects", file=sys.stderr)
        sys.exit(2)
    for i, item in enumerate(data, 1):
        if not isinstance(item, dict):
            print(f"ERROR: Element #{i} is not an object", file=sys.stderr)
            sys.exit(2)
    return data


def autofit_columns(ws, max_width=80, min_width=10):
    # Approximate Excel autofit by measuring string length (largest line if multi-line)
    for col in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col)
        max_len = 0
        for row in range(1, ws.max_row + 1):
            v = ws.cell(row=row, column=col).value
            if v is None:
                continue
            s = str(v)
            # consider multi-line
            for line in s.splitlines() or [s]:
                if len(line) > max_len:
                    max_len = len(line)
        width = max(min_width, min(max_width, max_len + 2))
        ws.column_dimensions[col_letter].width = width


def apply_borders(ws):
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = THIN_BORDER


def add_validation_dropdown(ws, start_row=2, end_row=50000):
    last_col = ws.max_column
    last_col_letter = get_column_letter(last_col)
    # Allow: Required, Not Required, and blank
    dv = DataValidation(type="list", formula1='"Required,Not Required"', allow_blank=True, showErrorMessage=True)
    dv.error = "Select one of: Required, Not Required (or leave blank)"
    ws.add_data_validation(dv)
    dv.ranges.append(f"{last_col_letter}{start_row}:{last_col_letter}{end_row}")


def write_testplan_sheet(wb, rows):
    ws = wb.active
    ws.title = "TestPlan"

    # Header
    for c_idx, header in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=c_idx, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(vertical="center")

    # Rows
    for r_idx, obj in enumerate(rows, start=2):
        for c_idx, header in enumerate(COLUMNS, start=1):
            if header == "Index":
                val = r_idx - 1
            elif header == NEW_COL_HEADER:
                val = ""  # default empty
            else:
                val = to_str(obj.get(header, ""))
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            wrap = header in WRAP_COLS
            cell.alignment = Alignment(wrap_text=wrap, vertical="top")

    # Freeze first row
    ws.freeze_panes = "A2"

    # Validation dropdown on last column
    add_validation_dropdown(ws)

    # Borders
    apply_borders(ws)

    # Autofit columns (approx)
    autofit_columns(ws)

    return ws


def add_raw_json_sheet(wb, data):
    raw = wb.create_sheet("RawJSON")
    # Store raw JSON for traceability (avoid any data loss concerns)
    raw.cell(row=1, column=1, value="raw_json")
    raw.cell(row=2, column=1, value=json.dumps(data, ensure_ascii=False, indent=2))
    # Hide the sheet to keep workbook clean
    raw.sheet_state = "hidden"


def main():
    parser = argparse.ArgumentParser(description="Generate TestPlan Excel from JSON")
    parser.add_argument("--json-path", required=True, help="Path to the Test Plan JSON file in the repo")
    parser.add_argument("--output-dir", required=True, help="Directory to write the Excel file to")
    args = parser.parse_args()

    rows = load_json_array(args.json_path)

    wb = Workbook()
    write_testplan_sheet(wb, rows)
    add_raw_json_sheet(wb, rows)

    # Ensure output directory exists
    os.makedirs(args.output_dir, exist_ok=True)
    ts = ist_now_str()
    filename = f"testplan_{ts}.xlsx"
    out_path = os.path.join(args.output_dir, filename)

    wb.save(out_path)
    # Print exact path for CI to capture
    print(out_path)


if __name__ == "__main__":
    main()
