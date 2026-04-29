#!/usr/bin/env python3
import argparse
import json
import os
from datetime import datetime
from zoneinfo import ZoneInfo
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

BLUE_FILL = PatternFill(fill_type="solid", fgColor="FF0070C0")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin")
)
HEADER_FONT = Font(bold=True)
CENTER = Alignment(horizontal="center", vertical="center")
TOP_LEFT = Alignment(horizontal="left", vertical="top")
TOP_CENTER = Alignment(horizontal="center", vertical="top")

WRAP_COLUMNS = {
    "Test Description",
    "Remarks",
    "Test Steps / Procedure",
    "Validation / Acceptance Criteria",
}


def load_rows_from_json(json_obj):
    # If explicit testcases array exists, prefer it
    if isinstance(json_obj, dict) and isinstance(json_obj.get("testcases"), list):
        rows = json_obj["testcases"]
    elif isinstance(json_obj, list):
        rows = json_obj
    elif isinstance(json_obj, dict):
        rows = [json_obj]
    else:
        raise ValueError("Unsupported JSON structure for tabular conversion")
    if not rows:
        raise ValueError("Empty JSON rows")
    # Ensure each row is a dict
    normalized = []
    for r in rows:
        if isinstance(r, dict):
            normalized.append(r)
        else:
            # Convert primitives to a dict with a generic column
            normalized.append({"value": r})
    return normalized


def union_headers_in_order(rows):
    seen = []
    seen_set = set()
    # first row order
    for k in rows[0].keys():
        seen.append(k)
        seen_set.add(k)
    # subsequent new keys in first-seen order
    for r in rows[1:]:
        for k in r.keys():
            if k not in seen_set:
                seen.append(k)
                seen_set.add(k)
    return seen


def cell_value(v):
    # Preserve primitives; serialize lists/dicts to JSON string
    if isinstance(v, (list, dict)):
        return json.dumps(v, ensure_ascii=False)
    return v


def is_numeric_column(header):
    return header.strip().lower() in {"index", "id", "no", "num", "number"}


def autofit_columns(ws):
    # approximate width by max display length
    col_max = {}
    for row in ws.iter_rows(values_only=True):
        for idx, v in enumerate(row, start=1):
            s = "" if v is None else str(v)
            col_max[idx] = max(col_max.get(idx, 0), len(s))
    # also include header row
    for idx, cell in enumerate(ws[1], start=1):
        col_max[idx] = max(col_max.get(idx, 0), len(str(cell.value) if cell.value is not None else ""))
    for idx, width in col_max.items():
        ws.column_dimensions[get_column_letter(idx)].width = min(max(width + 2, 10), 120)


def autofit_row_heights(ws):
    # Basic line-based height scaling for wrapped text
    base = 15
    for row in ws.iter_rows(min_row=2):
        max_lines = 1
        for c in row:
            v = "" if c.value is None else str(c.value)
            lines = v.count("\n") + 1
            if lines > max_lines:
                max_lines = lines
        ws.row_dimensions[c.row].height = base * max_lines


def apply_formatting(ws, headers):
    # Freeze top row
    ws.freeze_panes = "A2"

    # Header styles
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.fill = BLUE_FILL
        cell.border = THIN_BORDER

    # Data cells formatting
    for r in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
        for c in r:
            c.border = THIN_BORDER
            c.alignment = TOP_LEFT

    # Wrap text for specific columns
    header_to_col = {ws.cell(row=1, column=i+1).value: i+1 for i in range(len(headers))}
    for name in WRAP_COLUMNS:
        col = header_to_col.get(name)
        if col:
            for row_idx in range(2, ws.max_row + 1):
                ws.cell(row=row_idx, column=col).alignment = Alignment(wrapText=True, vertical="top", horizontal="left")

    # Numeric/index columns center aligned
    for i, h in enumerate(headers, start=1):
        if is_numeric_column(h):
            for row_idx in range(2, ws.max_row + 1):
                ws.cell(row=row_idx, column=i).alignment = TOP_CENTER

    # Autofit
    autofit_columns(ws)
    autofit_row_heights(ws)


def main():
    ap = argparse.ArgumentParser(description="Generate formatted Excel from JSON")
    ap.add_argument("--json-path", required=True)
    ap.add_argument("--output-dir", required=True)
    ap.add_argument("--ip-name", required=True)
    args = ap.parse_args()

    with open(args.json_path, "r", encoding="utf-8") as f:
        data = json.load(f)

    rows = load_rows_from_json(data)
    headers = union_headers_in_order(rows)

    wb = Workbook()
    ws = wb.active
    ws.title = "Data"

    # Write headers
    for i, h in enumerate(headers, start=1):
        ws.cell(row=1, column=i, value=h)

    # Write rows
    for r_idx, row in enumerate(rows, start=2):
        for c_idx, h in enumerate(headers, start=1):
            ws.cell(row=r_idx, column=c_idx, value=cell_value(row.get(h, "")))

    # Rename to TestPlan and format
    ws.title = "TestPlan"
    apply_formatting(ws, headers)

    # Ensure output dir exists
    os.makedirs(args.output_dir, exist_ok=True)

    # Compute IST timestamp
    ist_now = datetime.now(ZoneInfo("Asia/Kolkata"))
    ts = ist_now.strftime("%Y%m%d_%H%M%S")
    filename = f"{args.ip_name}_TestPlan_{ts}.xlsx"
    out_path = os.path.join(args.output_dir, filename)

    wb.save(out_path)
    print(f"Saved: {out_path}")

if __name__ == "__main__":
    main()
