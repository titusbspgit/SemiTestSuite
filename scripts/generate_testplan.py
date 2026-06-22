#!/usr/bin/env python3
# Ag-Emb-Testsuite-Excel-Generator Agent fallback script
# Converts JSON -> REAL .xlsx with required structure, validation, and formatting

import json
import sys
from pathlib import Path
from datetime import datetime
try:
    from zoneinfo import ZoneInfo  # Python 3.9+
except Exception:  # pragma: no cover
    ZoneInfo = None

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

INPUT_JSON_PATH = Path("Test_Output/GPIO/testplan_input.json")
OUTPUT_DIR = Path("Test_Output/GPIO")
SHEET_NAME = "TestPlan"
HEADERS = [
    "Index",
    "SS / Module",
    "Feature",
    "Test Case Name",
    "Test Description",
    "Speed",
    "Mode",
    "Remarks",
    "Test Steps / Procedure",
    "Impacted Registers",
    "Validation / Acceptance Criteria",
    "Gap Analysis",
    "Code Generation (Required / Not)",
]

WRAP_COLS = {
    5,   # Test Description
    9,   # Test Steps / Procedure
    11,  # Validation / Acceptance Criteria
}


def eprint(*args, **kwargs):
    print(*args, file=sys.stderr, **kwargs)


def ensure_ist_timestamp():
    if ZoneInfo is None:
        # Fallback: compute IST as UTC+5:30 without tz DB
        from datetime import timezone, timedelta
        ist = timezone(timedelta(hours=5, minutes=30))
        return datetime.now(ist).strftime("%Y%m%d_%H%M%S")
    else:
        ist = ZoneInfo("Asia/Kolkata")
        return datetime.now(ist).strftime("%Y%m%d_%H%M%S")


def cell_value(v):
    if v is None:
        return ""
    # Preserve complex structures as JSON strings deterministically
    if isinstance(v, (dict, list)):
        return json.dumps(v, ensure_ascii=False)
    return str(v)


def main():
    if not INPUT_JSON_PATH.exists():
        eprint(f"ERROR: Input JSON not found at {INPUT_JSON_PATH}.")
        eprint("Please add your aggregated Test Plan JSON array to this path and re-run.")
        sys.exit(1)

    raw = INPUT_JSON_PATH.read_text(encoding="utf-8")
    try:
        data = json.loads(raw)
    except json.JSONDecodeError as ex:
        eprint(f"ERROR: Invalid JSON: {ex}")
        sys.exit(1)

    if not isinstance(data, list):
        eprint("ERROR: json_data must be a JSON array of objects (each object is one row).")
        sys.exit(1)
    for i, row in enumerate(data, 1):
        if not isinstance(row, dict):
            eprint(f"ERROR: Element at index {i-1} is not an object: {type(row)}")
            sys.exit(1)

    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME

    # Header styles
    header_font = Font(bold=True)
    align_wrap = Alignment(wrapText=True, vertical="top")
    align_top = Alignment(vertical="top")
    thin = Side(style="thin")
    border_all = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Write headers
    for col_idx, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.alignment = align_top
        cell.border = border_all

    # Write rows
    for r_idx, row in enumerate(data, start=2):
        for c_idx, header in enumerate(HEADERS, start=1):
            if header == "Code Generation (Required / Not)":
                val = ""  # leave empty by default
            else:
                val = cell_value(row.get(header, ""))
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.alignment = (align_wrap if c_idx in WRAP_COLS else align_top)
            cell.border = border_all

    max_row = ws.max_row
    max_col = ws.max_column

    # Data validation dropdown for last column across all rows
    dv = DataValidation(type="list", formula1='"Required,Not Required"', allow_blank=True, showErrorMessage=True)
    ws.add_data_validation(dv)
    dv.add(f"M2:M1048576")  # Apply to all possible rows in column M (13th column)

    # Freeze first row
    ws.freeze_panes = "A2"

    # Autofit columns (approximation) with cap
    max_widths = [len(str(h)) for h in HEADERS]
    for r in range(2, max_row + 1):
        for c in range(1, max_col + 1):
            v = ws.cell(row=r, column=c).value
            if v is None:
                l = 0
            else:
                s = str(v)
                # For wrapped columns, use longest line length
                if c in WRAP_COLS:
                    l = max((len(line) for line in s.splitlines()), default=0)
                else:
                    l = len(s)
            if l > max_widths[c - 1]:
                max_widths[c - 1] = l

    for c in range(1, max_col + 1):
        width = min(max_widths[c - 1] + 2, 80)
        ws.column_dimensions[chr(64 + c)].width = width

    # Save workbook
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    ts = ensure_ist_timestamp()
    out_path = OUTPUT_DIR / f"testplan_{ts}.xlsx"
    wb.save(out_path)
    print(f"Generated: {out_path}")


if __name__ == "__main__":
    main()
