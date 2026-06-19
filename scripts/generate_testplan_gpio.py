#!/usr/bin/env python3
# coding: utf-8

import json
import os
from pathlib import Path
from typing import Any, Dict, List

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

# ==== Embedded JSON from ag_emb_testsuite_testplan_extract (use EXACTLY as provided) ====
RAW_JSON = r'''{
  "meta": {
    "ip_name": "GPIO",
    "repo": "titusbspgit/SemiTestSuite",
    "branch": "main",
    "source_folders": [
      "https://github.com/titusbspgit/SemiTestSuite/tree/main/TestRepo/gpio/gpio_reg_wr_rd_test",
      "https://github.com/titusbspgit/SemiTestSuite/tree/main/TestRepo/gpio/test_gpio_nedge_alternate_pads_en",
      "https://github.com/titusbspgit/SemiTestSuite/tree/main/TestRepo/gpio/test_gpio_negedge_intr_en",
      "https://github.com/titusbspgit/SemiTestSuite/tree/main/TestRepo/gpio/test_gpio_pedge_all_pads_en",
      "https://github.com/titusbspgit/SemiTestSuite/tree/main/TestRepo/gpio/test_gpio_pedge_alternate_pads_en"
    ]
  },
  "testcases": [
    {
      "test_id": "gpio_reg_wr_rd_test",
      "name": "gpio_reg_wr_rd_test",
      "title": "GPIO Reg Wr Rd Test",
      "description": "Enables GPIOA clock, configures PA0 as output, writes high/low to ODR and verifies the bit state.",
      "tags": [
        "GPIO"
      ],
      "priority": null,
      "owner": null,
      "parameters": [],
      "expected_results": "Prints: GPIO REG WR/RD TEST PASSED",
      "file_name": "program.c",
      "file_path": "TestRepo/gpio/gpio_reg_wr_rd_test/program.c",
      "file_github_url": "https://github.com/titusbspgit/SemiTestSuite/blob/main/TestRepo/gpio/gpio_reg_wr_rd_test/program.c",
      "source_folder_url": "https://github.com/titusbspgit/SemiTestSuite/tree/main/TestRepo/gpio/gpio_reg_wr_rd_test",
      "framework": "C"
    },
    {
      "test_id": "test_gpio_nedge_alternate_pads_en",
      "name": "test_gpio_nedge_alternate_pads_en",
      "title": "GPIO Negative Edge Enable on Alternate Pads",
      "description": "Enables EXTI falling-edge trigger on alternate pins {0,2,4,6} and verifies each bit is set.",
      "tags": [
        "GPIO"
      ],
      "priority": null,
      "owner": null,
      "parameters": [],
      "expected_results": "Prints: GPIO NEDGE ALT PADS TEST PASSED",
      "file_name": "program.c",
      "file_path": "TestRepo/gpio/test_gpio_nedge_alternate_pads_en/program.c",
      "file_github_url": "https://github.com/titusbspgit/SemiTestSuite/blob/main/TestRepo/gpio/test_gpio_nedge_alternate_pads_en/program.c",
      "source_folder_url": "https://github.com/titusbspgit/SemiTestSuite/tree/main/TestRepo/gpio/test_gpio_nedge_alternate_pads_en",
      "framework": "C"
    },
    {
      "test_id": "test_gpio_negedge_intr_en",
      "name": "test_gpio_negedge_intr_en",
      "title": "GPIO Negative Edge Interrupt Enable",
      "description": "Enables EXTI for PA1 with falling-edge trigger by configuring RCC, SYSCFG, and EXTI registers; verifies FTSR bit is set.",
      "tags": [
        "GPIO"
      ],
      "priority": null,
      "owner": null,
      "parameters": [],
      "expected_results": "Prints: GPIO NEG EDGE INTR TEST PASSED",
      "file_name": "program.c",
      "file_path": "TestRepo/gpio/test_gpio_negedge_intr_en/program.c",
      "file_github_url": "https://github.com/titusbspgit/SemiTestSuite/blob/main/TestRepo/gpio/test_gpio_negedge_intr_en/program.c",
      "source_folder_url": "https://github.com/titusbspgit/SemiTestSuite/tree/main/TestRepo/gpio/test_gpio_negedge_intr_en",
      "framework": "C"
    },
    {
      "test_id": "test_gpio_pedge_all_pads_en",
      "name": "test_gpio_pedge_all_pads_en",
      "title": "GPIO Positive Edge Enable on All Pads",
      "description": "Enables EXTI rising-edge trigger for pins 0 through 7 and verifies each bit is set.",
      "tags": [
        "GPIO"
      ],
      "priority": null,
      "owner": null,
      "parameters": [],
      "expected_results": "Prints: GPIO PEDGE ALL PADS TEST PASSED",
      "file_name": "program.c",
      "file_path": "TestRepo/gpio/test_gpio_pedge_all_pads_en/program.c",
      "file_github_url": "https://github.com/titusbspgit/SemiTestSuite/blob/main/TestRepo/gpio/test_gpio_pedge_all_pads_en/program.c",
      "source_folder_url": "https://github.com/titusbspgit/SemiTestSuite/tree/main/TestRepo/gpio/test_gpio_pedge_all_pads_en",
      "framework": "C"
    },
    {
      "test_id": "test_gpio_pedge_alternate_pads_en",
      "name": "test_gpio_pedge_alternate_pads_en",
      "title": "GPIO Positive Edge Enable on Alternate Pads",
      "description": "Enables EXTI rising-edge trigger on alternate pins {0,2,4,6} and verifies each bit is set.",
      "tags": [
        "GPIO"
      ],
      "priority": null,
      "owner": null,
      "parameters": [],
      "expected_results": "Prints: GPIO PEDGE ALT PADS TEST PASSED",
      "file_name": "program.c",
      "file_path": "TestRepo/gpio/test_gpio_pedge_alternate_pads_en/program.c",
      "file_github_url": "https://github.com/titusbspgit/SemiTestSuite/blob/main/TestRepo/gpio/test_gpio_pedge_alternate_pads_en/program.c",
      "source_folder_url": "https://github.com/titusbspgit/SemiTestSuite/tree/main/TestRepo/gpio/test_gpio_pedge_alternate_pads_en",
      "framework": "C"
    }
  ]
}'''

# Output path (relative to repo root)
OUTPUT_PATH = Path("Test_Output/GPIO/TestPlan.xlsx")


def main() -> None:
    data: Dict[str, Any] = json.loads(RAW_JSON)
    if not isinstance(data, dict) or "testcases" not in data or not isinstance(data["testcases"], list):
        raise SystemExit("Invalid input JSON: expected an object with 'testcases' array")

    testcases: List[Dict[str, Any]] = data["testcases"]

    wb = Workbook()
    ws = wb.active
    ws.title = "TestPlan"

    # Columns required for this export
    headers = [
        "test_id",
        "name",
        "title",
        "description",
        "tags",
        "priority",
        "owner",
        "parameters",
        "expected_results",
        "file_name",
        "file_path",
        "file_github_url",
        "source_folder_url",
        "framework",
        "Code Generation (Required / Not)",  # Extra column with validation
    ]

    # Write header
    ws.append(headers)
    header_font = Font(bold=True)
    for cell in ws[1]:
        cell.font = header_font

    # Freeze header row
    ws.freeze_panes = "A2"

    # Populate rows
    for tc in testcases:
        tags = tc.get("tags")
        tags_str = ", ".join(tags) if isinstance(tags, list) else (tags if isinstance(tags, str) else "")
        params = tc.get("parameters")
        if isinstance(params, list) and len(params) == 0:
            params_str = ""
        else:
            params_str = json.dumps(params, ensure_ascii=False) if params is not None else ""

        row = [
            tc.get("test_id") or "",
            tc.get("name") or "",
            tc.get("title") or "",
            tc.get("description") or "",
            tags_str,
            tc.get("priority") if tc.get("priority") is not None else "",
            tc.get("owner") if tc.get("owner") is not None else "",
            params_str,
            tc.get("expected_results") or "",
            tc.get("file_name") or "",
            tc.get("file_path") or "",
            tc.get("file_github_url") or "",
            tc.get("source_folder_url") or "",
            tc.get("framework") or "",
            "",  # Code Generation dropdown column (left empty by default)
        ]
        ws.append(row)

    # Data validation dropdown for the last column
    last_col_letter = ws.cell(row=1, column=len(headers)).column_letter
    dv = DataValidation(type="list", formula1='"Required,Not Required"', allow_blank=True, showDropDown=True)
    dv.error = "Select from the list: Required or Not Required"
    dv.prompt = "Choose if code generation is required"
    ws.add_data_validation(dv)
    dv.add(f"{last_col_letter}2:{last_col_letter}1000")  # apply to many rows

    # Text wrapping for long text columns: description and expected_results
    wrap_alignment = Alignment(wrap_text=True, vertical="top")
    desc_col_idx = headers.index("description") + 1
    exp_col_idx = headers.index("expected_results") + 1

    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    max_col_widths = [len(h) for h in headers]

    for r in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(headers)):
        for c_idx, cell in enumerate(r, start=1):
            # Borders
            cell.border = border
            # Wrap for specific columns
            if c_idx in (desc_col_idx, exp_col_idx):
                cell.alignment = wrap_alignment
            # Track widths
            val = cell.value
            text = str(val) if val is not None else ""
            if len(text) > max_col_widths[c_idx - 1]:
                max_col_widths[c_idx - 1] = min(len(text), 120)

    # Apply borders to header row as well and set alignment
    for cell in ws[1]:
        cell.border = border
        cell.alignment = Alignment(vertical="center")

    # Approximate autofit
    from openpyxl.utils import get_column_letter
    for idx, width in enumerate(max_col_widths, start=1):
        col_letter = get_column_letter(idx)
        # heuristic: width in Excel units
        ws.column_dimensions[col_letter].width = min(max(width + 2, 12), 80)

    # Ensure output directory exists
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)

    # Save workbook
    wb.save(str(OUTPUT_PATH))


if __name__ == "__main__":
    main()
