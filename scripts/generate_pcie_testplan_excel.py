#!/usr/bin/env python3
import json
import os
from pathlib import Path
from typing import List, Dict, Any

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment
except ImportError as e:
    raise SystemExit("openpyxl is required. Install with `pip install openpyxl`. Error: %s" % e)

DATA_JSON_PATH = Path("data/testplans/PCIE_testplan_consolidated.json")
OUTPUT_XLSX_PATH = Path("Test_Output/PCIE/TestPlan.xlsx")
SHEET_NAME = "PCIE_TestPlan"

HEADERS = [
    "Index",
    "SS / Module",
    "Feature",
    "Test Case Name",
    "Test Description",
    "Speed",
    "Mode",
    "Remarks",
    "Test Steps / Procedure",
    "Impacted Registers",
    "Validation / Acceptance Criteria",
    "Gap Analysis",
]


def derive_feature(name: str) -> str:
    if not name:
        return ""
    if name.endswith("_test"):
        return name[:-5]
    return name


def compact_remarks(tc: Dict[str, Any]) -> str:
    keep_keys = [
        "files", "path", "folder_url", "framework", "parameters",
        "tags", "priority", "type", "owner", "references", "id"
    ]
    data = {k: tc.get(k) for k in keep_keys if k in tc}
    try:
        return json.dumps(data, separators=(",", ":"), ensure_ascii=False)
    except Exception:
        return str(data)


def autosize_columns(ws):
    # Compute max width per column
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                val = cell.value
                if val is None:
                    continue
                length = len(str(val))
                if length > max_len:
                    max_len = length
            except Exception:
                pass
        # Cap width to avoid overly wide columns
        ws.column_dimensions[col_letter].width = min(80, max(10, max_len + 2))


def main() -> None:
    if not DATA_JSON_PATH.exists():
        raise SystemExit(f"Input JSON not found: {DATA_JSON_PATH}")

    with DATA_JSON_PATH.open("r", encoding="utf-8") as f:
        root = json.load(f)

    # Validate structure
    if not isinstance(root, dict) or "testcases" not in root or not isinstance(root["testcases"], list):
        raise SystemExit("Invalid input JSON: expected object with 'testcases' array")

    testcases: List[Dict[str, Any]] = root["testcases"]
    ip_name = root.get("ip_name", "")

    # Prepare workbook
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME

    # Header row
    ws.append(HEADERS)
    bold_font = Font(bold=True)
    for cell in ws[1]:
        cell.font = bold_font

    # Freeze top row
    ws.freeze_panes = "A2"

    wrap_alignment = Alignment(wrap_text=True, vertical="top")

    # Populate rows
    for i, tc in enumerate(testcases, start=1):
        steps_list = tc.get("steps") or []
        expected_list = tc.get("expected_results") or []
        row = [
            i,  # Index
            ip_name or "",  # SS / Module
            derive_feature(tc.get("name", "")),  # Feature
            tc.get("name", ""),  # Test Case Name
            tc.get("description", ""),  # Test Description
            "NA",  # Speed
            "NA",  # Mode
            compact_remarks(tc),  # Remarks
            "\n".join(str(s) for s in steps_list),  # Test Steps / Procedure
            "NA",  # Impacted Registers (not provided in consolidated schema)
            "\n".join(str(e) for e in expected_list),  # Validation / Acceptance Criteria
            "NA",  # Gap Analysis
        ]
        ws.append(row)

    # Apply wrap text to multi-line columns
    multiline_cols = {
        "H",  # Remarks
        "I",  # Test Steps / Procedure
        "K",  # Validation / Acceptance Criteria
    }
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            if cell.column_letter in multiline_cols:
                cell.alignment = wrap_alignment

    # Autosize columns
    autosize_columns(ws)

    # Ensure directory exists and save
    OUTPUT_XLSX_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_XLSX_PATH)
    print(f"Wrote Excel: {OUTPUT_XLSX_PATH}")


if __name__ == "__main__":
    main()
