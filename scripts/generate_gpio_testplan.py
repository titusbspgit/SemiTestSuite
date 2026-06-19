#!/usr/bin/env python3
# Ag-Emb-Testsuite-Excel-Generator Agent
# Purpose: Convert provided Test Plan JSON into a REAL .xlsx Excel file with formatting and data validation.
# Output: Test_Output/GPIO/GPIO_TestPlan.xlsx (fixed name per task-specific rules)
# Triggered via: workflow_dispatch or push to this script (CI will generate and commit Excel)

import os
import json
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

# ---- Source JSON (use exactly as provided; do not alter) ----
JSON_INPUT = r'''{
  "status": "SUCCESS",
  "data": [
    {
      "Index": "1",
      "SS / Module": "GPIO",
      "Feature": "Register write/read on GPIOA ODR (PA0)",
      "Test Case Name": "gpio_reg_wr_rd_test",
      "Test Description": "Validates GPIOA clock enable, PA0 output configuration, and write/read of ODR bit0 by setting high then low and checking readback.",
      "Speed": "NA",
      "Mode": "Polling",
      "Remarks": "Requires target with RCC and GPIOA at specified memory map; uses ITM for PASS/FAIL prints; program halts in an infinite loop after PASS.",
      "Test Steps / Procedure": "1. Enable AHB1 clock for GPIOA via RCC_AHB1ENR.\n2. Configure PA0 as output by updating GPIOA_MODER.\n3. Set PA0 high by writing GPIOA_ODR bit0.\n4. Read back GPIOA_ODR bit0 and verify it is set.\n5. Clear PA0 low by clearing GPIOA_ODR bit0.\n6. Read back GPIOA_ODR bit0 and verify it is cleared.",
      "Impacted Registers": "RCC_AHB1ENR, GPIOA_MODER, GPIOA_ODR",
      "Validation / Acceptance Criteria": "PASS if (GPIOA_ODR & (1<<0)) is true after setting high and (GPIOA_ODR & (1<<0)) is false after clearing low; otherwise prints FAIL and halts.",
      "Gap Analysis": "- No verification of actual pad state beyond ODR readback.\n- Hardcoded pin (PA0) and addresses; no parameterization.\n- No negative/error handling or cleanup."
    },
    {
      "Index": "2",
      "SS / Module": "GPIO",
      "Feature": "Falling-edge trigger enable on alternate EXTI lines",
      "Test Case Name": "test_gpio_nedge_alternate_pads_en",
      "Test Description": "Enables EXTI falling-edge trigger for alternate lines (0,2,4,6) and verifies configuration via readback.",
      "Speed": "NA",
      "Mode": "Polling",
      "Remarks": "Assumes EXTI block accessible and lines 0,2,4,6 are valid; uses ITM for PASS/FAIL prints; program halts after PASS.",
      "Test Steps / Procedure": "1. For pins 0, 2, 4, 6, set the corresponding bit in EXTI_FTSR.\n2. After each write, read EXTI_FTSR and verify the bit is set.\n3. If any verification fails, declare FAIL; otherwise declare PASS.\n4. Enter infinite loop after PASS.",
      "Impacted Registers": "EXTI_FTSR",
      "Validation / Acceptance Criteria": "PASS if for each of pins {0,2,4,6}, (EXTI_FTSR & (1<<pin)) evaluates true after enabling; otherwise prints FAIL and halts.",
      "Gap Analysis": "- No IMR configuration or source pin mapping validated.\n- Hardcoded pin list; no boundary or error cases.\n- Does not test actual interrupt/event occurrence."
    },
    {
      "Index": "3",
      "SS / Module": "GPIO",
      "Feature": "Negative-edge interrupt configuration on EXTI1 (PA1)",
      "Test Case Name": "test_gpio_negedge_intr_en",
      "Test Description": "Enables clocks, configures PA1 as input and maps to EXTI1, enables IMR and falling-edge trigger, then verifies EXTI1 FTSR bit.",
      "Speed": "NA",
      "Mode": "Polling",
      "Remarks": "Requires RCC AHB1/APB2 and SYSCFG access; assumes PA1 is mapped to EXTI1; uses ITM for PASS/FAIL prints; program halts after PASS.",
      "Test Steps / Procedure": "1. Enable GPIOA and SYSCFG clocks via RCC_AHB1ENR and RCC_APB2ENR.\n2. Configure PA1 as input via GPIOA_MODER.\n3. Map PA1 to EXTI1 by clearing SYSCFG_EXTICR1 field for EXTI1.\n4. Enable EXTI1 in EXTI_IMR and set falling-edge in EXTI_FTSR.\n5. Verify EXTI_FTSR bit1 is set for line 1.\n6. Declare PASS if set; otherwise FAIL and halt.",
      "Impacted Registers": "RCC_AHB1ENR, RCC_APB2ENR, GPIOA_MODER, SYSCFG_EXTICR1, EXTI_IMR, EXTI_FTSR",
      "Validation / Acceptance Criteria": "PASS if (EXTI_FTSR & (1<<1)) evaluates true after configuration; otherwise prints \"FAIL: NEG EDGE NOT ENABLED\" and halts.",
      "Gap Analysis": "- No NVIC/ISR configuration or interrupt servicing to confirm runtime behavior.\n- Limited to FTSR readback; IMR/mapping not independently verified beyond writes.\n- Hardcoded line (EXTI1/PA1) and addresses."
    },
    {
      "Index": "4",
      "SS / Module": "GPIO",
      "Feature": "Rising-edge trigger enable on all EXTI lines 0–7",
      "Test Case Name": "test_gpio_pedge_all_pads_en",
      "Test Description": "Enables EXTI rising-edge trigger for lines 0 through 7 and verifies each via readback.",
      "Speed": "NA",
      "Mode": "Polling",
      "Remarks": "Assumes EXTI lines 0–7 are present and writable; uses ITM for PASS/FAIL prints; program halts after PASS.",
      "Test Steps / Procedure": "1. For i from 0 to 7, set the corresponding bit in EXTI_RTSR.\n2. After each write, read EXTI_RTSR and verify the bit is set.\n3. If any verification fails, declare FAIL; otherwise declare PASS.\n4. Enter infinite loop after PASS.",
      "Impacted Registers": "EXTI_RTSR",
      "Validation / Acceptance Criteria": "PASS if for each i in [0..7], (EXTI_RTSR & (1<<i)) evaluates true after enabling; otherwise prints FAIL and halts.",
      "Gap Analysis": "- No IMR configuration or pin-to-line mapping validation.\n- Does not verify actual event generation; configuration-only.\n- Hardcoded line range (0–7); no bounds or error handling."
    },
    {
      "Index": "5",
      "SS / Module": "GPIO",
      "Feature": "Rising-edge trigger enable on alternate EXTI lines",
      "Test Case Name": "test_gpio_pedge_alternate_pads_en",
      "Test Description": "Enables EXTI rising-edge trigger for alternate lines (0,2,4,6) and verifies configuration via readback.",
      "Speed": "NA",
      "Mode": "Polling",
      "Remarks": "Assumes EXTI block accessible and lines 0,2,4,6 valid; uses ITM for PASS/FAIL prints; program halts after PASS.",
      "Test Steps / Procedure": "1. For pins 0, 2, 4, 6, set the corresponding bit in EXTI_RTSR.\n2. After each write, read EXTI_RTSR and verify the bit is set.\n3. If any verification fails, declare FAIL; otherwise declare PASS.\n4. Enter infinite loop after PASS.",
      "Impacted Registers": "EXTI_RTSR",
      "Validation / Acceptance Criteria": "PASS if for each of pins {0,2,4,6}, (EXTI_RTSR & (1<<pin)) evaluates true after enabling; otherwise prints FAIL and halts.",
      "Gap Analysis": "- Configuration-only; no end-to-end interrupt/event validation.\n- Hardcoded pin set; no negative or boundary testing.\n- No cleanup or restoration of default states."
    }
  ],
  "errors": []
}'''

# ---- Config ----
OUTPUT_PATH = os.path.join("Test_Output", "GPIO", "GPIO_TestPlan.xlsx")
SHEET_NAME = "TestPlan"
HEADERS = [
    "Index",
    "SS / Module",
    "Feature",
    "Test Case Name",
    "Test Description",
    "Speed",
    "Mode",
    "Remarks",
    "Test Steps / Procedure",
    "Impacted Registers",
    "Validation / Acceptance Criteria",
    "Gap Analysis",
    "Code Generation (Required / Not)",  # New column
]
WRAP_COLS = {
    "Test Description",
    "Test Steps / Procedure",
    "Validation / Acceptance Criteria",
}


def validate_json(raw: str):
    obj = json.loads(raw)
    if not isinstance(obj, dict):
        raise ValueError("Top-level JSON must be an object with keys 'status' and 'data'.")
    if obj.get("status") != "SUCCESS":
        raise ValueError("JSON status must be 'SUCCESS'.")
    data = obj.get("data")
    if not isinstance(data, list):
        raise ValueError("'data' must be a list of rows.")
    for i, row in enumerate(data, 1):
        if not isinstance(row, dict):
            raise ValueError(f"Row {i} is not an object.")
    return data


def build_workbook(rows):
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME

    # Header row
    for col_idx, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(vertical="center")

    # Data rows
    for r_idx, row in enumerate(rows, start=2):
        for c_idx, header in enumerate(HEADERS, start=1):
            if header == "Code Generation (Required / Not)":
                value = ""  # default empty
            else:
                value = row.get(header, "")
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            # Wrap alignment for certain columns
            if header in WRAP_COLS:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            else:
                cell.alignment = Alignment(vertical="top")

    # Freeze top row
    ws.freeze_panes = "A2"

    # Data validation for the last column (entire column from row 2 downward)
    last_col_idx = len(HEADERS)
    last_col_letter = get_column_letter(last_col_idx)
    dv = DataValidation(type="list", formula1='"Required,Not Required"', allow_blank=True, showErrorMessage=True)
    dv.error = "Select a value from the dropdown: Required or Not Required (or leave blank)."
    ws.add_data_validation(dv)
    dv.ranges.append(f"{last_col_letter}2:{last_col_letter}1048576")  # apply to all potential rows

    # Apply thin borders to all populated cells
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    max_row = ws.max_row
    max_col = ws.max_column
    for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
        for cell in row:
            cell.border = border

    # Autofit columns (approximation)
    col_widths = {}
    for col_idx in range(1, max_col + 1):
        header = HEADERS[col_idx - 1]
        max_len = len(str(header)) if header else 0
        for row_idx in range(2, max_row + 1):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val is None:
                continue
            # Use a conservative width estimation
            val_str = str(val)
            line_max = max((len(line) for line in val_str.splitlines()), default=0)
            max_len = max(max_len, line_max)
        # Base width scaling; cap to reasonable limits
        base = max_len * 1.1 + 2
        if header in WRAP_COLS:
            base = max(base, 60)  # ensure readability for wrapped text
        base = min(max(base, 10), 120)
        col_widths[col_idx] = base

    for col_idx, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    return wb


def main():
    rows = validate_json(JSON_INPUT)

    # Ensure output directory exists
    out_dir = os.path.dirname(OUTPUT_PATH)
    if out_dir and not os.path.exists(out_dir):
        os.makedirs(out_dir, exist_ok=True)

    wb = build_workbook(rows)
    wb.save(OUTPUT_PATH)
    print(f"Saved Excel to: {OUTPUT_PATH}")


if __name__ == "__main__":
    main()

# CI trigger