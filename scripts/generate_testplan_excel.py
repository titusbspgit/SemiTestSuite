#!/usr/bin/env python3
import argparse
import json
import os
from datetime import datetime, timedelta, timezone
from typing import List, Dict

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation


REQUIRED_COLUMNS = [
    "Index",
    "SS / Module",
    "Feature",
    "Test Case Name",
    "Test Description",
    "Speed",
    "Mode",
    "Remarks",
    "Test Steps / Procedure",
    "Impacted Registers",
    "Validation / Acceptance Criteria",
    "Gap Analysis",
    "Code Generation (Required / Not)",
]

WRAP_COLUMNS = {
    "Test Description",
    "Test Steps / Procedure",
    "Validation / Acceptance Criteria",
}


def ist_timestamp() -> str:
    # IST is UTC+05:30; avoid dependency on zoneinfo for portability
    ist = timezone(timedelta(hours=5, minutes=30))
    now_ist = datetime.now(ist)
    return now_ist.strftime("%Y%m%d_%H%M%S")


def validate_json(data) -> List[Dict]:
    if not isinstance(data, list):
        raise ValueError("json_data must be a JSON array")
    for i, item in enumerate(data, 1):
        if not isinstance(item, dict):
            raise ValueError(f"Element at index {i} is not an object")
    return data


def ensure_dir(path: str):
    os.makedirs(path, exist_ok=True)


def compute_column_widths(ws, header_to_index):
    # Approximate autofit by max string length per column, capped
    max_len = {h: len(h) for h in header_to_index.keys()}
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        for h, idx in header_to_index.items():
            val = row[idx - 1]
            if val is None:
                continue
            length = len(str(val))
            if length > max_len[h]:
                max_len[h] = length
    for h, idx in header_to_index.items():
        # Wider default for wrapped columns
        if h in WRAP_COLUMNS:
            width = min(max(max_len[h] * 0.9, 40), 100)
        else:
            width = min(max(max_len[h] * 0.9, 12), 60)
        ws.column_dimensions[chr(64 + idx)].width = width


def apply_borders(ws):
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = border


def main():
    parser = argparse.ArgumentParser(description="Generate TestPlan Excel from JSON")
    parser.add_argument("--input", required=True, help="Path to input JSON file (array of objects)")
    parser.add_argument("--output-dir", required=True, help="Directory to write Excel file into")
    args = parser.parse_args()

    with open(args.input, "r", encoding="utf-8") as f:
        data = json.load(f)

    rows = validate_json(data)

    wb = Workbook()
    ws = wb.active
    ws.title = "TestPlan"

    # Freeze header row
    ws.freeze_panes = "A2"

    # Write header
    for col_idx, header in enumerate(REQUIRED_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)

    # Map JSON to required columns
    for r_idx, item in enumerate(rows, start=2):
        for c_idx, header in enumerate(REQUIRED_COLUMNS, start=1):
            if header == "Code Generation (Required / Not)":
                value = ""
            else:
                value = item.get(header, "")
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            # Wrap selected columns
            if header in WRAP_COLUMNS:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            else:
                cell.alignment = Alignment(vertical="top")

    # Data Validation for last column
    last_col_idx = REQUIRED_COLUMNS.index("Code Generation (Required / Not)") + 1
    dv = DataValidation(type="list", formula1='"Required,Not Required"', allow_blank=True, showDropDown=True)
    dv.error = "Select a value from the list"
    dv.errorTitle = "Invalid Input"
    dv.prompt = "Choose 'Required' or 'Not Required' (or leave blank)"
    dv.promptTitle = "Code Generation"
    start_row = 2
    end_row = ws.max_row if ws.max_row >= 2 else 2
    dv_range = f"{chr(64 + last_col_idx)}{start_row}:{chr(64 + last_col_idx)}{end_row}"
    dv.add(dv_range)
    ws.add_data_validation(dv)

    # Apply borders
    apply_borders(ws)

    # Approximate autofit
    header_to_index = {h: i + 1 for i, h in enumerate(REQUIRED_COLUMNS)}
    compute_column_widths(ws, header_to_index)

    # Add RawData sheet to preserve all JSON fields exactly
    raw = wb.create_sheet("RawData")
    # Determine key order by first appearance across rows
    keys_order = []
    seen = set()
    for obj in rows:
        for k in obj.keys():
            if k not in seen:
                seen.add(k)
                keys_order.append(k)
    # Header
    for i, k in enumerate(keys_order, start=1):
        cell = raw.cell(row=1, column=i, value=k)
        cell.font = Font(bold=True)
    # Rows
    for r, obj in enumerate(rows, start=2):
        for i, k in enumerate(keys_order, start=1):
            cell = raw.cell(row=r, column=i, value=obj.get(k, ""))
            cell.alignment = Alignment(vertical="top")
    apply_borders(raw)
    # Basic widths for raw
    for i in range(1, raw.max_column + 1):
        raw.column_dimensions[chr(64 + i)].width = 24

    ensure_dir(args.output_dir)
    ts = ist_timestamp()
    filename = f"testplan_{ts}.xlsx"
    out_path = os.path.join(args.output_dir, filename)
    wb.save(out_path)

    print(out_path)
    # Expose path to later workflow steps if running in GitHub Actions
    gha = os.environ.get("GITHUB_OUTPUT")
    if gha:
        with open(gha, "a", encoding="utf-8") as g:
            g.write(f"excel_path={out_path}\n")


if __name__ == "__main__":
    main()
