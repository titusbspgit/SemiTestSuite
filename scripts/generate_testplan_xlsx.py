#!/usr/bin/env python3
# Purpose: Convert Test Plan JSON into a REAL .xlsx with required formatting and validation
# Requirements implemented:
# - Sheet name: "TestPlan"
# - Exact headers and order
# - Last column: "Code Generation (Required / Not)" with dropdown (Required, Not Required, allow blank)
# - Bold header, freeze first row, wrap long text columns, borders on all used cells, autofit columns
# - File name: testplan_<YYYYMMDD_HHMMSS IST>.xlsx

import argparse
import json
import os
from datetime import datetime

try:
    import pytz
except ImportError:  # Fallback if pytz isn't available; Actions installs it
    pytz = None

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

HEADERS = [
    "Index",
    "SS / Module",
    "Feature",
    "Test Case Name",
    "Test Description",
    "Speed",
    "Mode",
    "Remarks",
    "Test Steps / Procedure",
    "Impacted Registers",
    "Validation / Acceptance Criteria",
    "Gap Analysis",
    "Code Generation (Required / Not)",
]

# Columns requiring text wrap by 1-based Excel index (E, I, K)
WRAP_COL_IDXS = {5, 9, 11}


def ist_timestamp():
    tz = None
    if pytz is not None:
        tz = pytz.timezone("Asia/Kolkata")
    if tz:
        return datetime.now(tz).strftime("%Y%m%d_%H%M%S")
    # Fallback to naive local time if pytz absent (GH Actions installs pytz)
    return datetime.now().strftime("%Y%m%d_%H%M%S")


def load_json(json_path):
    with open(json_path, "r", encoding="utf-8") as f:
        data = json.load(f)
    if not isinstance(data, list):
        raise ValueError("json_data must be a JSON array of objects (each object = one row)")
    for i, row in enumerate(data, start=1):
        if not isinstance(row, dict):
            raise ValueError(f"Element at index {i-1} is not an object; found {type(row)}")
    return data


def sanitize_value(v):
    if v is None:
        return ""
    if isinstance(v, (str, int, float)):
        return v
    # Preserve exact JSON for complex types
    return json.dumps(v, ensure_ascii=False)


def autofit_columns(ws, max_col, max_row):
    # Compute an approximate best-fit width per column
    # Using a simple heuristic based on max text length
    from openpyxl.utils import get_column_letter

    MIN_WIDTH = 10
    MAX_WIDTH = 80

    for col_idx in range(1, max_col + 1):
        max_length = 0
        # include header row
        header_val = ws.cell(row=1, column=col_idx).value
        if header_val is not None:
            max_length = max(max_length, len(str(header_val)))
        for row_idx in range(2, max_row + 1):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val is None:
                continue
            s = str(val)
            # Treat newlines as extra width contributors
            s_no_newlines = s.replace("\n", " ")
            max_length = max(max_length, len(s_no_newlines))
        # A heuristic factor to convert chars to Excel width units
        width = min(MAX_WIDTH, max(MIN_WIDTH, max_length * 1.1))
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def apply_borders_and_alignment(ws, max_col, max_row):
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = border
            if r == 1:
                # Header alignment: center vertically, wrap off by default
                cell.alignment = Alignment(vertical="center")
            else:
                # Wrap where required, vertical top to aid readability
                wrap = c in WRAP_COL_IDXS
                cell.alignment = Alignment(wrap_text=wrap, vertical="top")


def build_workbook(rows):
    wb = Workbook()
    ws = wb.active
    ws.title = "TestPlan"

    # Header row
    for col_idx, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)

    # Data rows
    for r_idx, row in enumerate(rows, start=2):
        for c_idx, header in enumerate(HEADERS, start=1):
            if header == "Code Generation (Required / Not)":
                val = ""  # default empty
            else:
                val = sanitize_value(row.get(header, ""))
            ws.cell(row=r_idx, column=c_idx, value=val)

    # Freeze first row
    ws.freeze_panes = "A2"

    # Data validation for last column (M): Required / Not Required, allow blank
    last_col_letter = "M"  # 13th column
    dv = DataValidation(type="list", formula1='"Required,Not Required"', allow_blank=True, showErrorMessage=True)
    dv.error = "Select a value from the list or leave blank."
    dv.errorTitle = "Invalid Selection"
    ws.add_data_validation(dv)
    dv.add(f"{last_col_letter}2:{last_col_letter}1048576")  # apply to entire column starting row 2

    # Autofit columns based on content
    max_row = ws.max_row
    max_col = len(HEADERS)
    autofit_columns(ws, max_col, max_row)

    # Apply borders and alignment (including wrapping on selected columns)
    apply_borders_and_alignment(ws, max_col, max_row)

    return wb


def main():
    parser = argparse.ArgumentParser(description="Generate TestPlan Excel from JSON")
    parser.add_argument("--json", dest="json_path", default="data/testplan.json", help="Path to JSON file")
    parser.add_argument("--outdir", dest="outdir", default="Test_Output/GPIO", help="Output directory for Excel")
    args = parser.parse_args()

    rows = load_json(args.json_path)

    wb = build_workbook(rows)

    os.makedirs(args.outdir, exist_ok=True)
    ts = ist_timestamp()
    out_path = os.path.join(args.outdir, f"testplan_{ts}.xlsx")
    wb.save(out_path)
    print(f"Generated Excel: {out_path}")


if __name__ == "__main__":
    main()
