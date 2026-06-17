#!/usr/bin/env python3
# Generates a REAL Excel (.xlsx) Test Plan from embedded JSON
# Sheet name: "TestPlan"
# Applies formatting, borders, frozen header, wrapped text, and a dropdown validation

import os
from datetime import datetime
from zoneinfo import ZoneInfo

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# Repository output directory (created if missing)
OUTPUT_DIR = "Test_Output/PCIE"

# Columns definition (order is mandatory). The last column is new and must be empty with dropdown.
HEADERS = [
    "Index",
    "SS / Module",
    "Feature",
    "Test Case Name",
    "Test Description",
    "Speed",
    "Mode",
    "Remarks",
    "Test Steps / Procedure",
    "Impacted Registers",
    "Validation / Acceptance Criteria",
    "Gap Analysis",
    "Code Generation (Required / Not)",  # NEW LAST COLUMN
]

# Embedded JSON data (preserved as provided; no transformation)
JSON_DATA = [
    {
        "Index": "1",
        "SS / Module": "NA",
        "Feature": "NA",
        "Test Case Name": "NA",
        "Test Description": "NA",
        "Speed": "NA",
        "Mode": "NA",
        "Remarks": "NA",
        "Test Steps / Procedure": "NA",
        "Impacted Registers": "NA",
        "Validation / Acceptance Criteria": "NA",
        "Gap Analysis": "NA"
    }
]

# Columns requiring text wrap
WRAP_HEADERS = {
    "Test Description",
    "Test Steps / Procedure",
    "Validation / Acceptance Criteria",
}


def main():
    # Ensure output directory exists
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    # IST timestamp for filename
    ist = ZoneInfo("Asia/Kolkata")
    ts = datetime.now(ist).strftime("%Y%m%d_%H%M%S")
    out_path = os.path.join(OUTPUT_DIR, f"testplan_{ts}.xlsx")

    wb = Workbook()
    ws = wb.active
    ws.title = "TestPlan"

    # Header row
    for col_idx, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)

    # Data rows (preserve provided values; last column stays blank)
    for r_idx, row in enumerate(JSON_DATA, start=2):
        for c_idx, header in enumerate(HEADERS, start=1):
            if header == "Code Generation (Required / Not)":
                value = None  # keep empty
            else:
                # Use exact header key to preserve mapping; default empty if missing
                value = row.get(header, "")
            cell = ws.cell(row=r_idx, column=c_idx, value=value)

    # Freeze top row
    ws.freeze_panes = "A2"

    # Data Validation for last column (dropdown: Required / Not Required; allow blank)
    last_col_index = len(HEADERS)
    last_col_letter = get_column_letter(last_col_index)
    dv = DataValidation(type="list", formula1='"Required,Not Required"', allow_blank=True)
    dv.error = "Select either 'Required' or 'Not Required' or leave blank."
    dv.errorTitle = "Invalid Selection"
    dv.prompt = "Choose if code generation is required."
    dv.promptTitle = "Code Generation"
    ws.add_data_validation(dv)
    dv.add(f"{last_col_letter}2:{last_col_letter}1048576")  # Apply to all possible rows

    # Alignment (wrap text for specified columns)
    wrap_alignment = Alignment(wrap_text=True, vertical="top")
    normal_alignment = Alignment(vertical="top")

    wrap_indices = {HEADERS.index(h) + 1 for h in WRAP_HEADERS if h in HEADERS}

    max_row = ws.max_row
    max_col = ws.max_column

    # Apply borders and alignment to all cells
    thin = Side(border_style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = border
            if r == 1:
                # header already bold; keep default alignment top
                cell.alignment = Alignment(vertical="top")
            else:
                if c in wrap_indices:
                    cell.alignment = wrap_alignment
                else:
                    cell.alignment = normal_alignment

    # Auto-fit columns (approximate by content length)
    # Set a reasonable min/max width to keep sheet readable
    min_width = 10
    max_width = 60
    wrap_default_width = 40

    for c in range(1, max_col + 1):
        header = HEADERS[c - 1]
        col_letter = get_column_letter(c)
        max_len = len(str(header))
        for r in range(2, max_row + 1):
            val = ws.cell(row=r, column=c).value
            if val is None:
                ln = 0
            else:
                s = str(val)
                ln = len(s)
            if ln > max_len:
                max_len = ln
        if c in wrap_indices:
            width = min(max(wrap_default_width, min_width), max_width)
        else:
            width = min(max(max_len + 2, min_width), max_width)
        ws.column_dimensions[col_letter].width = width

    # Save workbook as REAL .xlsx
    wb.save(out_path)

    print(out_path)


if __name__ == "__main__":
    main()
